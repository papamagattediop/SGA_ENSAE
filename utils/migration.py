# ============================================================
#  SGA ENSAE — utils/migration.py
#  Import initial depuis Excel vers la base SQL
#  Python 3.11 · SQLAlchemy 2.0.30
# ============================================================

import sys
import os

# Ajouter le dossier racine du projet au path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import io
import base64
from datetime import datetime
from database import SessionLocal
from models import (
    Filiere, Classe, ResponsableClasse, User, Etudiant, RoleEnum,
    MigrationLog, StatutMigrationEnum
)
from auth import hash_password


# ============================================================
#  UTILITAIRES
# ============================================================

def log_migration(fichier: str, statut: str, details: str):
    """Enregistre un log de migration dans la base."""
    db = SessionLocal()
    try:
        db.add(MigrationLog(
            fichier     = fichier,
            date_import = datetime.now(),
            statut      = StatutMigrationEnum.succes if statut == "succes" else StatutMigrationEnum.erreur,
            details     = details
        ))
        db.commit()
    finally:
        db.close()


def parse_excel(contents: str) -> pd.ExcelFile:
    """Decode un fichier Excel base64 en ExcelFile pandas."""
    _, content_string = contents.split(",")
    decoded = base64.b64decode(content_string)
    return pd.ExcelFile(io.BytesIO(decoded))


# ============================================================
#  MIGRATION FILIERES
# ============================================================

def migrate_filieres(df: pd.DataFrame) -> tuple:
    """
    Colonnes attendues : Code, Libelle, Duree_ans
    Logique : si le code existe déjà → ignorer (pas de doublon)
    """
    db   = SessionLocal()
    nb   = 0
    err  = []
    warn = []
    try:
        codes_existants = {f.code.upper() for f in db.query(Filiere).all()}

        for idx, row in df.iterrows():
            try:
                code = str(row.get("Code", "")).strip().upper()
                if not code or code == "NAN" or code.startswith("#"):
                    continue
                if code in codes_existants:
                    warn.append(f"Filiere '{code}' deja en base — ignoree.")
                    continue
                libelle   = str(row.get("Libelle", "")).strip()
                duree_ans = int(row.get("Duree_ans", 3))
                db.add(Filiere(code=code, libelle=libelle, duree_ans=duree_ans))
                codes_existants.add(code)
                nb += 1
            except Exception as e:
                err.append(f"Filiere ligne {idx + 2} : {str(e)}")
        db.commit()
    except Exception as e:
        db.rollback()
        err.append(f"Erreur globale filieres : {str(e)}")
    finally:
        db.close()
    return nb, err, warn


# ============================================================
#  MIGRATION CLASSES
# ============================================================

def migrate_classes(df: pd.DataFrame) -> tuple:
    """
    Colonnes attendues : Nom, Code_Filiere, Niveau, Is_Commune, Annee_Scolaire
    """
    db   = SessionLocal()
    nb   = 0
    err  = []
    warn = []
    try:
        fil_map    = {f.code.upper(): f.id for f in db.query(Filiere).all()}
        existantes = {c.nom.strip().lower() for c in db.query(Classe).all()}

        for idx, row in df.iterrows():
            try:
                nom = str(row.get("Nom", "")).strip()
                if not nom or nom.startswith("#"):
                    continue

                code_fil = str(row.get("Code_Filiere", "")).strip().upper()
                if code_fil not in fil_map:
                    err.append(f"Filiere '{code_fil}' introuvable pour classe '{nom}'.")
                    continue

                annee = str(row.get("Annee_Scolaire", "")).strip()

                if nom.lower() in existantes:
                    warn.append(f"Classe '{nom}' deja en base — ignoree.")
                    continue

                db.add(Classe(
                    nom            = nom,
                    filiere_id     = fil_map[code_fil],
                    niveau         = int(row.get("Niveau", 1)),
                    is_commune     = bool(int(row.get("Is_Commune", 0))),
                    annee_scolaire = annee
                ))
                existantes.add(nom.lower())
                nb += 1
            except Exception as e:
                err.append(f"Classe ligne {idx + 2} : {str(e)}")
        db.commit()
    except Exception as e:
        db.rollback()
        err.append(f"Erreur globale classes : {str(e)}")
    finally:
        db.close()
    return nb, err, warn


# ============================================================
#  MIGRATION ETUDIANTS
# ============================================================

def migrate_etudiants(df: pd.DataFrame) -> tuple:
    """
    Colonnes attendues :
    Matricule, Nom, Prenom, Email, Date_Naissance,
    Nom_Classe, Annee_Scolaire, Filiere_Origine
    Mot_de_passe (optionnel — défaut = matricule)

    MODIFIÉ :
      - Lecture de Mot_de_passe (colonne optionnelle)
      - Si absent ou vide → mot de passe = matricule
      - Envoi email de bienvenue après création du compte
    """
    db   = SessionLocal()
    nb   = 0
    err  = []
    warn = []
    try:
        mats_existants   = {e.matricule for e in db.query(Etudiant).all()}
        emails_existants = {u.email.lower() for u in db.query(User).all()}
        cls_map = {
            c.nom.strip().lower(): c.id
            for c in db.query(Classe).all()
        }
        # Charger les noms de classes pour l'email de bienvenue
        cls_nom_map = {
            c.id: c.nom
            for c in db.query(Classe).all()
        }

        for idx, row in df.iterrows():
            try:
                matricule = str(row.get("Matricule", "")).strip()
                email     = str(row.get("Email", "")).strip().lower()

                if not matricule or matricule.upper() == "NAN" or matricule.startswith("#"):
                    continue

                if matricule in mats_existants:
                    warn.append(f"Matricule '{matricule}' deja en base — ignore.")
                    continue
                if email in emails_existants:
                    warn.append(f"Email '{email}' deja en base — ignore.")
                    continue

                nom_classe = str(row.get("Nom_Classe", "")).strip()
                annee      = str(row.get("Annee_Scolaire", "")).strip()
                cls_key    = nom_classe.strip().lower()
                if cls_key not in cls_map:
                    err.append(f"Classe '{nom_classe}' introuvable pour '{matricule}'.")
                    continue

                # Parsing date de naissance
                date_naissance = None
                raw_date = str(row.get("Date_Naissance", "")).strip()
                if raw_date and raw_date.upper() != "NAN":
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                        try:
                            date_naissance = datetime.strptime(raw_date, fmt).date()
                            break
                        except ValueError:
                            continue

                # ── MODIFIÉ : lecture Mot_de_passe avec défaut = matricule ──
                mdp_brut = str(row.get("Mot_de_passe", "")).strip()
                if not mdp_brut or mdp_brut.upper() == "NAN":
                    mdp_brut = matricule  # défaut = matricule

                prenom = str(row.get("Prenom", "")).strip()
                nom    = str(row.get("Nom",    "")).strip()

                user = User(
                    nom           = nom.upper(),
                    prenom        = prenom,
                    email         = email,
                    password_hash = hash_password(mdp_brut),
                    role          = RoleEnum.eleve,
                    is_active     = True
                )
                db.add(user)
                db.flush()

                classe_id = cls_map[cls_key]
                db.add(Etudiant(
                    user_id         = user.id,
                    matricule       = matricule,
                    date_naissance  = date_naissance,
                    classe_id       = classe_id,
                    filiere_origine = str(row.get("Filiere_Origine", "")).strip(),
                    annee_scolaire  = annee
                ))
                mats_existants.add(matricule)
                emails_existants.add(email)
                nb += 1

                # ── MODIFIÉ : email de bienvenue après création ──────────
                # Envoyé en dehors du bloc try/except principal pour ne pas
                # bloquer la migration si l'envoi échoue
                nom_classe_affiche = cls_nom_map.get(classe_id, nom_classe)
                _envoyer_bienvenue_etudiant(
                    email         = email,
                    prenom        = prenom,
                    nom           = nom,
                    mot_de_passe  = mdp_brut,
                    classe        = nom_classe_affiche,
                    warn          = warn,
                )

            except Exception as e:
                err.append(f"Etudiant ligne {idx + 2} : {str(e)}")

        db.commit()
    except Exception as e:
        db.rollback()
        err.append(f"Erreur globale etudiants : {str(e)}")
    finally:
        db.close()
    return nb, err, warn


def _envoyer_bienvenue_etudiant(
    email: str, prenom: str, nom: str,
    mot_de_passe: str, classe: str, warn: list
) -> None:
    """
    Tente d'envoyer l'email de bienvenue à un étudiant nouvellement créé.
    En cas d'échec, ajoute un avertissement (non bloquant).
    """
    try:
        from utils.mailer import email_bienvenue_etudiant
        ok, err_msg = email_bienvenue_etudiant(
            to           = email,
            prenom       = prenom,
            nom          = nom,
            email_connexion = email,
            mot_de_passe = mot_de_passe,
            classe       = classe,
        )
        if not ok:
            warn.append(f"Email bienvenue non envoyé à '{email}' : {err_msg}")
    except Exception as e:
        warn.append(f"Email bienvenue non envoyé à '{email}' : {str(e)}")


# ============================================================
#  MIGRATION RESPONSABLES (filiere + classe)
# ============================================================

def migrate_responsables(df: pd.DataFrame, role_cible: str) -> tuple:
    """
    Colonnes :
    Resp_Filieres  : Nom | Prenom | Email | Mot_de_passe | Code_Filiere
    Resp_Classes   : Nom | Prenom | Email | Mot_de_passe | Nom_Classe | Type (titulaire/suppleant)

    Logique :
      - Email existant → promotion au bon rôle (pas de doublon)
      - Email nouveau  → création compte + affectation
      - Mot_de_passe optionnel si email existant
    """
    from models import ResponsableFiliere, ResponsableClasse

    db   = SessionLocal()
    nb   = 0
    err  = []
    warn = []

    try:
        fil_map   = {f.code.upper(): f.id for f in db.query(Filiere).all()}
        cls_map   = {c.nom.strip().lower(): c.id for c in db.query(Classe).all()}
        role_enum = RoleEnum.resp_filiere if role_cible == "resp_filiere" else RoleEnum.resp_classe

        for idx, row in df.iterrows():
            try:
                nom    = str(row.get("Nom",          "")).strip()
                prenom = str(row.get("Prenom",       "")).strip()
                email  = str(row.get("Email",        "")).strip().lower()
                mdp    = str(row.get("Mot_de_passe", "")).strip()

                if not email or email == "nan" or email.startswith("#"):
                    continue
                if not nom or nom.startswith("#"):
                    continue

                user = db.query(User).filter(User.email == email).first()
                if user:
                    user.role      = role_enum
                    user.is_active = True
                    if nom:    user.nom    = nom.upper()
                    if prenom: user.prenom = prenom
                    if mdp and len(mdp) >= 6:
                        user.password_hash = hash_password(mdp)
                    warn.append(f"'{email}' existe déjà → promu {role_cible}.")
                else:
                    if not mdp or len(mdp) < 6:
                        err.append(f"Mot de passe manquant/trop court pour '{email}' (min 6 car.).")
                        continue
                    user = User(
                        nom           = nom.upper(),
                        prenom        = prenom,
                        email         = email,
                        password_hash = hash_password(mdp),
                        role          = role_enum,
                        is_active     = True
                    )
                    db.add(user)
                    db.flush()

                # ── Affecter à la filière ou à la classe ────────────────
                if role_cible == "resp_filiere":
                    code_fil = str(row.get("Code_Filiere", "")).strip().upper()
                    if code_fil.startswith("#") or not code_fil or code_fil == "NAN":
                        continue
                    if code_fil not in fil_map:
                        err.append(f"Filière '{code_fil}' introuvable pour '{email}'.")
                        continue
                    old = db.query(ResponsableFiliere).filter(
                        ResponsableFiliere.user_id == user.id
                    ).first()
                    if old:
                        db.delete(old)
                    db.add(ResponsableFiliere(user_id=user.id, filiere_id=fil_map[code_fil]))

                else:  # resp_classe
                    nom_cls  = str(row.get("Nom_Classe", "")).strip()
                    type_del = str(row.get("Type", "titulaire")).strip().lower()
                    if nom_cls.startswith("#") or not nom_cls or nom_cls == "NAN":
                        continue
                    cls_key = nom_cls.lower()
                    if cls_key not in cls_map:
                        err.append(f"Classe '{nom_cls}' introuvable pour '{email}'.")
                        continue
                    est_tit = (type_del == "titulaire")
                    cls_id  = cls_map[cls_key]
                    old = db.query(ResponsableClasse).filter(
                        ResponsableClasse.classe_id     == cls_id,
                        ResponsableClasse.est_titulaire == est_tit
                    ).first()
                    if old:
                        db.delete(old)
                    db.add(ResponsableClasse(
                        user_id       = user.id,
                        classe_id     = cls_id,
                        est_titulaire = est_tit
                    ))

                nb += 1

            except Exception as e:
                err.append(f"Ligne {idx + 2} : {str(e)}")

        db.commit()

    except Exception as e:
        db.rollback()
        err.append(f"Erreur globale : {str(e)}")
    finally:
        db.close()

    return nb, err, warn


# ============================================================
#  MIGRATION DELEGUES (par matricule)
# ============================================================

def migrate_delegues(df, db) -> tuple:
    """
    Importe les délégués (titulaire + suppléant) depuis la feuille Delegues.
    Colonnes : Nom_Classe | Matricule_Titulaire | Matricule_Suppleant
    """
    from models import ResponsableClasse

    nb   = 0
    err  = []
    warn = []

    cols_req = ["Nom_Classe", "Matricule_Titulaire", "Matricule_Suppleant"]
    missing  = [c for c in cols_req if c not in df.columns]
    if missing:
        err.append(f"Colonnes manquantes : {', '.join(missing)}")
        return 0, err, warn

    for _, row in df.iterrows():
        nom_classe = str(row.get("Nom_Classe",            "")).strip()
        mat_tit    = str(row.get("Matricule_Titulaire",   "")).strip()
        mat_sup    = str(row.get("Matricule_Suppleant",   "")).strip()

        if not nom_classe:
            continue

        classe = db.query(Classe).filter(Classe.nom == nom_classe).first()
        if not classe:
            err.append(f"Classe '{nom_classe}' introuvable.")
            continue

        for mat, est_tit in [(mat_tit, True), (mat_sup, False)]:
            if not mat or mat == "nan":
                continue
            etu = db.query(Etudiant).filter(Etudiant.matricule == mat).first()
            if not etu:
                err.append(f"Matricule '{mat}' introuvable.")
                continue
            old = db.query(ResponsableClasse).filter(
                ResponsableClasse.classe_id     == classe.id,
                ResponsableClasse.est_titulaire == est_tit
            ).first()
            if old:
                db.delete(old)
            db.add(ResponsableClasse(
                user_id       = etu.user_id,
                classe_id     = classe.id,
                est_titulaire = est_tit
            ))
        nb += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        err.append(f"Erreur commit delegues : {str(e)}")

    return nb, err, warn


# ============================================================
#  MIGRATION COMPLETE DEPUIS UN FICHIER EXCEL
# ============================================================

def migrate_from_excel(contents: str, filename: str) -> dict:
    """
    Import complet depuis un fichier Excel multi-feuilles.

    Ordre recommandé :
    1. Filieres       : Code, Libelle, Duree_ans
    2. Classes        : Nom, Code_Filiere, Niveau, Is_Commune, Annee_Scolaire
    3. Etudiants      : Matricule, Nom, Prenom, Email, Date_Naissance,
                        Nom_Classe, Annee_Scolaire, Filiere_Origine, Mot_de_passe (optionnel)
    4. Resp_Filieres  : Nom, Prenom, Email, Mot_de_passe, Code_Filiere
    5. Resp_Classes   : Nom, Prenom, Email, Mot_de_passe, Nom_Classe, Type
    6. Delegues       : Nom_Classe, Matricule_Titulaire, Matricule_Suppleant
    """
    resultats      = {}
    toutes_erreurs = []

    try:
        xl = parse_excel(contents)

        # 1. Filières
        if "Filieres" in xl.sheet_names:
            nb, err, warn = migrate_filieres(xl.parse("Filieres"))
            resultats["Filieres"] = {"nb": nb, "erreurs": err, "avertissements": warn}
            toutes_erreurs.extend(err)

        # 2. Classes
        if "Classes" in xl.sheet_names:
            nb, err, warn = migrate_classes(xl.parse("Classes"))
            resultats["Classes"] = {"nb": nb, "erreurs": err, "avertissements": warn}
            toutes_erreurs.extend(err)

        # 3. Étudiants
        if "Etudiants" in xl.sheet_names:
            nb, err, warn = migrate_etudiants(xl.parse("Etudiants"))
            resultats["Etudiants"] = {"nb": nb, "erreurs": err, "avertissements": warn}
            toutes_erreurs.extend(err)

        # 4. Responsables de filières
        if "Resp_Filieres" in xl.sheet_names:
            nb, err, warn = migrate_responsables(xl.parse("Resp_Filieres"), "resp_filiere")
            resultats["Resp_Filieres"] = {"nb": nb, "erreurs": err, "avertissements": warn}
            toutes_erreurs.extend(err)

        # 5. Responsables de classes (délégués avec compte)
        if "Resp_Classes" in xl.sheet_names:
            nb, err, warn = migrate_responsables(xl.parse("Resp_Classes"), "resp_classe")
            resultats["Resp_Classes"] = {"nb": nb, "erreurs": err, "avertissements": warn}
            toutes_erreurs.extend(err)

        # 6. Délégués par matricule (sans création de compte)
        if "Delegues" in xl.sheet_names:
            db_session = SessionLocal()
            try:
                nb, err, warn = migrate_delegues(xl.parse("Delegues"), db_session)
            finally:
                db_session.close()
            resultats["Delegues"] = {"nb": nb, "erreurs": err, "avertissements": warn}
            toutes_erreurs.extend(err)

        # Statut global
        erreurs_fatales = [
            e for e in toutes_erreurs
            if not any(x in e for x in ["deja en base", "deja utilise",
                                         "existe deja", "→ promu"])
        ]
        statut  = "succes" if not erreurs_fatales else "erreur"
        details = f"{sum(r['nb'] for r in resultats.values())} enregistrement(s) importé(s). "
        if toutes_erreurs:
            details += " | ".join(toutes_erreurs[:3])
        log_migration(filename, statut, details)

    except Exception as e:
        log_migration(filename, "erreur", str(e))
        resultats["global"] = {"nb": 0, "erreurs": [f"Erreur fatale : {str(e)}"],
                               "avertissements": []}

    return resultats


# ============================================================
#  GENERATION TEMPLATE MIGRATION EXCEL
# ============================================================

def generate_migration_template() -> bytes:
    """
    Génère un template Excel vide avec en-têtes et lignes d'exemple grisées.
    L'utilisateur doit REMPLACER les lignes d'exemple par ses données.
    Les lignes commençant par '#' sont ignorées lors de l'import.

    MODIFIÉ :
      - Feuille Etudiants : ajout colonne Mot_de_passe (optionnelle)
      - Feuille Instructions : correction ISE 3 ans (était 5 ans)
                               + doc colonne Mot_de_passe + email bienvenue
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import openpyxl

    wb  = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Styles communs ─────────────────────────────────────────
    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    hdr_font    = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    exemple_font= Font(color="AAAAAA", italic=True, name="Calibri", size=10)
    exemple_fill= PatternFill("solid", fgColor="F5F5F5")
    border      = Border(
        left  =Side(style="thin", color="DDDDDD"),
        right =Side(style="thin", color="DDDDDD"),
        top   =Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD")
    )
    centre = Alignment(horizontal="center", vertical="center")
    gauche = Alignment(horizontal="left",   vertical="center")

    # ── Feuille Filieres ──────────────────────────────────────
    ws1 = wb.create_sheet("Filieres")
    ws1.sheet_properties.tabColor = "003580"
    headers_fil = ["Code", "Libelle", "Duree_ans"]
    for col, h in enumerate(headers_fil, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hdr_fill("003580"), hdr_font, centre, border
    # MODIFIÉ : ISE = 3 ans (corrigé de 5 ans)
    exemples_fil = [
        ["# ISEP",    "# Ingenieur Statisticien Economiste Preparatoire", "# 2"],
        ["# ISE",     "# Ingenieur Statisticien Economiste",              "# 3"],
        ["# AS",      "# Analyste Statisticien",                         "# 3"],
        ["# MASTERS", "# Masters Specialises",                           "# 1"],
    ]
    for r, row in enumerate(exemples_fil, 2):
        for col, val in enumerate(row, 1):
            c = ws1.cell(row=r, column=col, value=val)
            c.font, c.fill, c.alignment, c.border = exemple_font, exemple_fill, gauche, border
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 55
    ws1.column_dimensions["C"].width = 12
    ws1.row_dimensions[1].height = 20

    # ── Feuille Classes ───────────────────────────────────────
    ws2 = wb.create_sheet("Classes")
    ws2.sheet_properties.tabColor = "006B3F"
    headers_cls = ["Nom", "Code_Filiere", "Niveau", "Is_Commune", "Annee_Scolaire"]
    for col, h in enumerate(headers_cls, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hdr_fill("006B3F"), hdr_font, centre, border
    exemples_cls = [
        ["# ISEP 1",     "# ISEP", "# 1", "# 0", "# 2024-2025"],
        ["# ISE Math 1", "# ISE",  "# 1", "# 0", "# 2024-2025"],
        ["# ISE 2",      "# ISE",  "# 2", "# 1", "# 2024-2025"],
        ["# AS1",        "# AS",   "# 1", "# 0", "# 2024-2025"],
    ]
    for r, row in enumerate(exemples_cls, 2):
        for col, val in enumerate(row, 1):
            c = ws2.cell(row=r, column=col, value=val)
            c.font, c.fill, c.alignment, c.border = exemple_font, exemple_fill, gauche, border
    for col_letter, width in zip(["A","B","C","D","E"], [20, 15, 10, 12, 14]):
        ws2.column_dimensions[col_letter].width = width
    ws2.row_dimensions[1].height = 20

    # ── Feuille Etudiants — MODIFIÉ : ajout Mot_de_passe ──────
    ws3 = wb.create_sheet("Etudiants")
    ws3.sheet_properties.tabColor = "F5A623"
    headers_etu = [
        "Matricule", "Nom", "Prenom", "Email",
        "Date_Naissance", "Nom_Classe", "Annee_Scolaire",
        "Filiere_Origine", "Mot_de_passe"        # ← NOUVEAU
    ]
    for col, h in enumerate(headers_etu, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hdr_fill("F5A623"), hdr_font, centre, border

    # Note explicative sur Mot_de_passe dans la cellule I1 (col 9)
    note_mdp = ws3.cell(row=1, column=10,
        value="⚠ Mot_de_passe optionnel. Si vide → défaut = Matricule. "
              "Un email de bienvenue avec les identifiants est envoyé automatiquement.")
    note_mdp.font = Font(color="EF4444", italic=True, size=9, name="Calibri")
    ws3.column_dimensions["J"].width = 70

    exemples_etu = [
        ["# ENSAE-2024-001", "# DIALLO", "# Amadou", "# adiallo@ensae.sn",
         "# 15/09/2002", "# ISEP 1", "# 2024-2025", "# ISEP", "# "],
        ["# ENSAE-2024-002", "# NDIAYE", "# Fatou",  "# fndiaye@ensae.sn",
         "# 20/03/2001", "# AS1",    "# 2024-2025", "# AS",   "# monMdpPerso"],
    ]
    for r, row in enumerate(exemples_etu, 2):
        for col, val in enumerate(row, 1):
            c = ws3.cell(row=r, column=col, value=val)
            c.font, c.fill, c.alignment, c.border = exemple_font, exemple_fill, gauche, border
    for col_letter, width in zip(["A","B","C","D","E","F","G","H","I"],
                                  [18, 15, 15, 25, 16, 16, 14, 15, 18]):
        ws3.column_dimensions[col_letter].width = width
    ws3.row_dimensions[1].height = 20

    # ── Feuille Delegues ──────────────────────────────────────
    ws4 = wb.create_sheet("Delegues")
    ws4.sheet_properties.tabColor = "8B5CF6"
    headers_del = ["Nom_Classe", "Matricule_Titulaire", "Matricule_Suppleant"]
    for col, h in enumerate(headers_del, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hdr_fill("8B5CF6"), hdr_font, centre, border
    exemples_del = [
        ["# AS1",    "# ENSAE-2024-001", "# ENSAE-2024-002"],
        ["# ISEP 1", "# ENSAE-2024-003", "# ENSAE-2024-004"],
    ]
    for r, row in enumerate(exemples_del, 2):
        for col, val in enumerate(row, 1):
            c = ws4.cell(row=r, column=col, value=val)
            c.font, c.fill, c.alignment, c.border = exemple_font, exemple_fill, gauche, border
    for col_letter, width in zip(["A","B","C"], [20, 22, 22]):
        ws4.column_dimensions[col_letter].width = width
    ws4.row_dimensions[1].height = 20

    # ── Feuille Resp. Filiere ─────────────────────────────────
    ws5 = wb.create_sheet("Resp_Filieres")
    ws5.sheet_properties.tabColor = "0EA5E9"
    headers_rf = ["Nom", "Prenom", "Email", "Mot_de_passe", "Code_Filiere"]
    for col, h in enumerate(headers_rf, 1):
        c = ws5.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hdr_fill("0EA5E9"), hdr_font, centre, border
    note_rf = ws5.cell(row=1, column=6,
        value="⚠ Mot_de_passe : min 6 caractères. Laissez vide si compte déjà existant.")
    note_rf.font = Font(color="EF4444", italic=True, size=9, name="Calibri")
    ws5.column_dimensions["F"].width = 55

    exemples_rf = [
        ["# Diallo", "# Ibrahima", "# idiallo@ensae.sn", "# pass123", "# ISE"],
        ["# Ndiaye", "# Aminata",  "# andiaye@ensae.sn", "# pass456", "# AS"],
    ]
    for r, row in enumerate(exemples_rf, 2):
        for col, val in enumerate(row, 1):
            c = ws5.cell(row=r, column=col, value=val)
            c.font, c.fill, c.alignment, c.border = exemple_font, exemple_fill, gauche, border
    for col_letter, width in zip(["A","B","C","D","E"], [15, 15, 28, 15, 14]):
        ws5.column_dimensions[col_letter].width = width
    ws5.row_dimensions[1].height = 20

    # ── Feuille Resp. Classe (Délégués) ───────────────────────
    ws6 = wb.create_sheet("Resp_Classes")
    ws6.sheet_properties.tabColor = "F59E0B"
    headers_rc = ["Nom", "Prenom", "Email", "Mot_de_passe", "Nom_Classe", "Type"]
    for col, h in enumerate(headers_rc, 1):
        c = ws6.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hdr_fill("F59E0B"), hdr_font, centre, border
    note_rc = ws6.cell(row=1, column=7,
        value="⚠ Type : 'titulaire' ou 'suppleant'. Email existant = promotion automatique.")
    note_rc.font = Font(color="EF4444", italic=True, size=9, name="Calibri")
    ws6.column_dimensions["G"].width = 55

    exemples_rc = [
        ["# Sall",  "# Marietou", "# sallmarietou@ensae.sn", "# pass123", "# AS3",    "# titulaire"],
        ["# Diop",  "# Magatte",  "# diopmagatte@ensae.sn",  "# pass456", "# AS3",    "# suppleant"],
        ["# Faye",  "# Ameth",    "# fayeameth@ensae.sn",    "# pass789", "# ISEP 1", "# titulaire"],
    ]
    for r, row in enumerate(exemples_rc, 2):
        for col, val in enumerate(row, 1):
            c = ws6.cell(row=r, column=col, value=val)
            c.font, c.fill, c.alignment, c.border = exemple_font, exemple_fill, gauche, border
    for col_letter, width in zip(["A","B","C","D","E","F"], [15, 15, 28, 15, 16, 13]):
        ws6.column_dimensions[col_letter].width = width
    ws6.row_dimensions[1].height = 20

    # ── Feuille Instructions — MODIFIÉ : ISE 3 ans + Mot_de_passe + email ──
    ws7 = wb.create_sheet("⚠ Instructions")
    ws7.sheet_properties.tabColor = "EF4444"
    ws7["A1"] = "INSTRUCTIONS — TEMPLATE DE MIGRATION SGA ENSAE"
    ws7["A1"].font = Font(bold=True, size=13, color="EF4444", name="Montserrat")

    instructions = [
        ("A3",  "RÈGLE GÉNÉRALE"),
        ("A4",  "→ Les lignes grisées commençant par '#' sont des EXEMPLES — supprimez-les avant d'importer."),
        ("A5",  "→ Ne modifiez JAMAIS les noms des colonnes (ligne 1)."),
        ("A6",  "→ L'ordre d'import est important : Filieres → Classes → Etudiants → Resp → Delegues."),

        ("A8",  "FEUILLE : Filieres"),
        ("A9",  "→ Ne remplir que si vos filières ne sont pas encore en base."),
        ("A10", "→ Code : identifiant court unique (ex: ISE, AS, ISEP, MASTERS)."),
        # MODIFIÉ : correction des durées — ISE = 3 ans, ISEP = 2 ans
        ("A11", "→ Durées officielles ENSAE : ISEP = 2 ans | ISE = 3 ans | AS = 3 ans | MASTERS = 1 an."),

        ("A13", "FEUILLE : Classes"),
        ("A14", "→ Ne remplir que si vos classes ne sont pas encore en base."),
        ("A15", "→ Is_Commune : 0 = classe normale, 1 = classe commune/partagée entre filières."),

        ("A17", "FEUILLE : Etudiants"),
        ("A18", "→ Une ligne par étudiant."),
        # MODIFIÉ : doc Mot_de_passe + email bienvenue
        ("A19", "→ Mot_de_passe : OPTIONNEL. Si vide → mot de passe par défaut = Matricule."),
        ("A20", "→ Un email de bienvenue avec les identifiants de connexion est envoyé automatiquement à chaque étudiant importé."),
        ("A21", "→ Date_Naissance : format JJ/MM/AAAA (ex: 15/09/2002) ou AAAA-MM-JJ."),

        ("A23", "FEUILLE : Resp_Filieres"),
        ("A24", "→ Crée un compte responsable de filière avec les identifiants fournis."),
        ("A25", "→ Si l'email existe déjà en base → promotion automatique au rôle resp_filiere."),
        ("A26", "→ Mot_de_passe : obligatoire pour un nouveau compte, optionnel si email existant."),

        ("A28", "FEUILLE : Resp_Classes (Délégués)"),
        ("A29", "→ Type : 'titulaire' ou 'suppleant' (un de chaque par classe)."),
        ("A30", "→ Si l'email est celui d'un étudiant existant → promotion automatique."),

        ("A32", "FEUILLE : Delegues"),
        ("A33", "→ Alternative à Resp_Classes : désigner par matricule étudiant existant."),
        ("A34", "→ Matricule_Titulaire et Matricule_Suppleant doivent déjà être importés."),
    ]

    for cell_ref, text in instructions:
        ws7[cell_ref] = text
        is_header = not text.startswith("→")
        ws7[cell_ref].font = Font(
            size=10 if not is_header else 11,
            bold=is_header,
            name="Calibri",
            color="111827" if not is_header else "003580"
        )
    ws7.column_dimensions["A"].width = 95

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ============================================================
#  SCRIPT STANDALONE
# ============================================================

if __name__ == "__main__":
    """
    Utilisation en ligne de commande :
    python utils/migration.py --template
    python utils/migration.py --import fichier.xlsx
    """

    if len(sys.argv) < 2:
        print("Usage : python migration.py --template | --import <fichier.xlsx>")
        sys.exit(1)

    if sys.argv[1] == "--template":
        data = generate_migration_template()
        with open("template_migration_ensae.xlsx", "wb") as f:
            f.write(data)
        print("[OK] Template genere : template_migration_ensae.xlsx")

    elif sys.argv[1] == "--import" and len(sys.argv) >= 3:
        filepath = sys.argv[2]
        print(f"[...] Import depuis {filepath}")
        with open(filepath, "rb") as f:
            raw = f.read()
        b64 = ("data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
               + base64.b64encode(raw).decode())
        resultats = migrate_from_excel(b64, filepath)
        for feuille, res in resultats.items():
            print(f"  {feuille} : {res['nb']} importe(s)")
            for e in res["erreurs"]:
                print(f"    [!] {e}")
        print("[OK] Migration terminee.")