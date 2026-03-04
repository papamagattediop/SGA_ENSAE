# ============================================================
#  SGA ENSAE — pages/etudiants.py
#  Module 5 : Gestion des etudiants et notes
#  Python 3.11 · Dash 2.17.0
# ============================================================

import dash
from dash import html, dcc, callback, Input, Output, State, ctx
import dash_bootstrap_components as dbc
from auth import require_auth
from database import SessionLocal
from models import Etudiant, User, Classe, Note, Presence, Seance, Module, UE, RoleEnum
from sqlalchemy import func
import pandas as pd
import io
import base64

dash.register_page(__name__, path="/etudiants", title="SGA ENSAE — Etudiants")

BLEU = "#003580"
VERT = "#006B3F"
OR   = "#F5A623"


# ============================================================
#  UTILITAIRES DB
# ============================================================

def get_classes():
    db = SessionLocal()
    try:
        return [{"label": c.nom, "value": c.id} for c in db.query(Classe).all()]
    finally:
        db.close()

def get_etudiants(classe_id=None, search=None):
    db = SessionLocal()
    try:
        q = db.query(Etudiant).join(User)
        if classe_id:
            q = q.filter(Etudiant.classe_id == classe_id)
        if search:
            q = q.filter(
                (User.nom.ilike(f"%{search}%")) |
                (User.prenom.ilike(f"%{search}%")) |
                (Etudiant.matricule.ilike(f"%{search}%"))
            )
        etudiants = q.order_by(User.nom, User.prenom).all()
        result = []
        for e in etudiants:
            # Taux assiduite
            total_seances = db.query(Presence).filter(Presence.etudiant_id == e.id).count()
            presences     = db.query(Presence).filter(
                Presence.etudiant_id == e.id, Presence.present == True
            ).count()
            taux = round((presences / total_seances) * 100) if total_seances > 0 else 100

            # Moyenne generale
            notes = db.query(Note).filter(Note.etudiant_id == e.id).all()
            moy   = round(sum(n.note for n in notes) / len(notes), 2) if notes else None

            result.append({
                "id"           : e.id,
                "matricule"    : e.matricule,
                "nom"          : e.user.nom if e.user else "",
                "prenom"       : e.user.prenom if e.user else "",
                "email"        : e.user.email if e.user else "",
                "classe"       : e.classe.nom if e.classe else "-",
                "classe_id"    : e.classe_id,
                "taux_assiduite": taux,
                "moyenne"      : moy,
                "nb_absences"  : total_seances - presences,
            })
        return result
    finally:
        db.close()

def get_etudiant_detail(etudiant_id):
    db = SessionLocal()
    try:
        e = db.query(Etudiant).filter(Etudiant.id == etudiant_id).first()
        if not e:
            return None

        # Notes par module
        notes_data = []
        modules = db.query(Module).filter(Module.classe_id == e.classe_id).all()
        for m in modules:
            notes_m = db.query(Note).filter(
                Note.etudiant_id == e.id,
                Note.module_id == m.id
            ).order_by(Note.numero).all()
            note1 = next((n.note for n in notes_m if n.numero == 1), None)
            note2 = next((n.note for n in notes_m if n.numero == 2), None)
            if note1 is not None and note2 is not None:
                moy_mod = round((note1 + note2) / 2, 2)
            elif note1 is not None:
                moy_mod = note1
            else:
                moy_mod = None
            notes_data.append({
                "module"    : m.libelle,
                "code"      : m.code,
                "ue"        : m.ue.libelle if m.ue else "-",
                "coef"      : m.coefficient,
                "note1"     : note1,
                "note2"     : note2,
                "moyenne"   : moy_mod,
                "module_id" : m.id,
            })

        # Absences
        absences = db.query(Presence).filter(
            Presence.etudiant_id == e.id, Presence.present == False
        ).all()

        return {
            "id"         : e.id,
            "matricule"  : e.matricule,
            "nom"        : e.user.nom if e.user else "",
            "prenom"     : e.user.prenom if e.user else "",
            "email"      : e.user.email if e.user else "",
            "classe"     : e.classe.nom if e.classe else "-",
            "naissance"  : e.date_naissance.strftime("%d/%m/%Y") if e.date_naissance else "-",
            "notes"      : notes_data,
            "absences"   : len(absences),
        }
    finally:
        db.close()

def get_modules_for_template(classe_id):
    db = SessionLocal()
    try:
        return db.query(Module).filter(Module.classe_id == classe_id).all()
    finally:
        db.close()

def generate_notes_template(classe_id) -> bytes:
    """Genere un template Excel pour la saisie des notes."""
    db = SessionLocal()
    try:
        etudiants = db.query(Etudiant).join(User).filter(
            Etudiant.classe_id == classe_id
        ).order_by(User.nom).all()
        modules = db.query(Module).filter(Module.classe_id == classe_id).all()

        rows = []
        for e in etudiants:
            row = {
                "ID_Etudiant": e.id,
                "Matricule"  : e.matricule,
                "Nom"        : e.user.nom if e.user else "",
                "Prenom"     : e.user.prenom if e.user else "",
            }
            for m in modules:
                row[f"{m.code}_Note1"] = ""
                row[f"{m.code}_Note2"] = ""
            rows.append(row)

        df = pd.DataFrame(rows)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Notes", index=False)
            wb = writer.book
            from openpyxl.styles import PatternFill, Font, Alignment
            ws = writer.sheets["Notes"]
            fill   = PatternFill("solid", fgColor="003580")
            font   = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill      = fill
                cell.font      = font
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 18

            # Feuille info modules
            df_mod = pd.DataFrame([
                {"Code": m.code, "Module": m.libelle,
                 "UE": m.ue.libelle if m.ue else "-",
                 "Coefficient": m.coefficient,
                 "Volume_h": m.volume_horaire}
                for m in modules
            ])
            df_mod.to_excel(writer, sheet_name="Info_Modules", index=False)

        buffer.seek(0)
        return buffer.read()
    finally:
        db.close()

def import_notes_from_excel(contents: str):
    """Importe les notes depuis un fichier Excel."""
    content_type, content_string = contents.split(",")
    decoded = base64.b64decode(content_string)
    df  = pd.read_excel(io.BytesIO(decoded), sheet_name="Notes")
    db  = SessionLocal()
    nb  = 0
    err = []
    try:
        for _, row in df.iterrows():
            etudiant_id = int(row["ID_Etudiant"])
            for col in df.columns:
                if "_Note" not in col:
                    continue
                parts      = col.rsplit("_Note", 1)
                module_code = parts[0]
                numero      = int(parts[1])
                val         = row[col]
                if pd.isna(val) or val == "":
                    continue
                note_val = float(val)
                if not (0 <= note_val <= 20):
                    err.append(f"Note hors plage (0-20) : {col} etudiant {etudiant_id}")
                    continue
                module = db.query(Module).filter(Module.code == module_code).first()
                if not module:
                    err.append(f"Module '{module_code}' introuvable.")
                    continue
                from models import TypeEvalEnum
                type_eval = TypeEvalEnum.devoir if numero == 1 else TypeEvalEnum.examen
                existing  = db.query(Note).filter(
                    Note.etudiant_id == etudiant_id,
                    Note.module_id   == module.id,
                    Note.numero      == numero
                ).first()
                if existing:
                    existing.note = note_val
                else:
                    db.add(Note(
                        etudiant_id=etudiant_id,
                        module_id=module.id,
                        note=note_val,
                        type_eval=type_eval,
                        numero=numero
                    ))
                nb += 1
        db.commit()
        return nb, err
    except Exception as e:
        db.rollback()
        err.append(f"Erreur globale : {str(e)}")
        return nb, err
    finally:
        db.close()


# ============================================================
#  COMPOSANTS UI
# ============================================================

def label_style():
    return {
        "fontWeight": "600", "fontSize": "0.82rem", "color": "#374151",
        "marginBottom": "6px", "display": "block", "fontFamily": "'Inter', sans-serif"
    }

def input_style():
    return {
        "borderRadius": "8px", "border": "1.5px solid #d1d5db",
        "fontSize": "0.875rem", "fontFamily": "'Inter', sans-serif"
    }

def field(label_text, component):
    return html.Div(style={"marginBottom": "14px"}, children=[
        html.Label(label_text, style=label_style()),
        component
    ])

def btn_primary(label, btn_id, color=BLEU):
    return html.Button(label, id=btn_id, n_clicks=0, style={
        "background": color, "color": "#ffffff", "border": "none",
        "borderRadius": "8px", "padding": "9px 18px",
        "fontFamily": "'Montserrat', sans-serif", "fontWeight": "600",
        "fontSize": "0.82rem", "cursor": "pointer"
    })

def btn_outline(label, btn_id):
    return html.Button(label, id=btn_id, n_clicks=0, style={
        "background": "transparent", "color": "#6b7280",
        "border": "1px solid #d1d5db", "borderRadius": "8px",
        "padding": "9px 18px", "fontFamily": "'Inter', sans-serif",
        "cursor": "pointer", "marginLeft": "8px"
    })

def taux_badge(taux):
    color = VERT if taux >= 80 else OR if taux >= 60 else "#ef4444"
    return html.Span(f"{taux}%", style={
        "background": f"{color}15", "color": color,
        "padding": "2px 8px", "borderRadius": "999px",
        "fontSize": "0.72rem", "fontWeight": "700",
        "fontFamily": "'Inter', sans-serif"
    })

def moy_badge(moy):
    if moy is None:
        return html.Span("-", style={"color": "#9ca3af", "fontSize": "0.82rem",
                                      "fontFamily": "'Inter', sans-serif"})
    color = VERT if moy >= 12 else OR if moy >= 10 else "#ef4444"
    return html.Span(f"{moy}/20", style={
        "background": f"{color}15", "color": color,
        "padding": "2px 8px", "borderRadius": "999px",
        "fontSize": "0.72rem", "fontWeight": "700",
        "fontFamily": "'Inter', sans-serif"
    })


# ============================================================
#  LAYOUT
# ============================================================

layout = html.Div([
    dcc.Store(id="etudiant-selectionne"),
    dcc.Store(id="etudiants-refresh"),
    dcc.Download(id="download-notes-template"),

    # -- Retour + En-tete --
    html.Div(style={"marginBottom": "24px"}, children=[
        html.A(href="/dashboard", style={
            "textDecoration": "none", "display": "inline-flex",
            "alignItems": "center", "gap": "6px", "marginBottom": "12px",
            "color": "#6b7280", "fontFamily": "'Inter', sans-serif",
            "fontSize": "0.82rem", "fontWeight": "500"
        }, children=[
            html.Span("arrow_back", className="material-symbols-outlined",
                      style={"fontSize": "18px", "verticalAlign": "middle"}),
            "Retour au tableau de bord"
        ]),
        html.H4("Gestion des Etudiants", style={
            "fontFamily": "'Montserrat', sans-serif",
            "fontWeight": "800", "color": BLEU, "margin": "0"
        }),
        html.P("Fiches etudiants, notes et assiduite", style={
            "color": "#6b7280", "fontFamily": "'Inter', sans-serif", "margin": "4px 0 0"
        })
    ]),

    # -- Barre filtres + actions --
    html.Div(style={
        "background": "#ffffff", "borderRadius": "12px",
        "padding": "16px 24px", "border": "1px solid #e5e7eb",
        "marginBottom": "20px", "display": "flex",
        "alignItems": "center", "gap": "12px", "flexWrap": "wrap"
    }, children=[
        html.Div(style={"flex": "1", "minWidth": "160px"}, children=[
            dcc.Dropdown(id="etudiants-filtre-classe",
                         placeholder="Filtrer par classe...",
                         style={"fontFamily": "'Inter', sans-serif", "fontSize": "0.875rem"})
        ]),
        html.Div(style={"flex": "2", "minWidth": "200px"}, children=[
            dbc.Input(id="etudiants-search", placeholder="Rechercher nom, prenom ou matricule...",
                      style=input_style())
        ]),
        # Notes workflow
        html.Div(style={"display": "flex", "gap": "8px", "alignItems": "center"}, children=[
            html.Div(style={"minWidth": "180px"}, children=[
                dcc.Dropdown(id="notes-classe-select",
                             placeholder="Classe pour notes...",
                             style={"fontFamily": "'Inter', sans-serif", "fontSize": "0.875rem"})
            ]),
            html.Button(
                [html.Span("download", className="material-symbols-outlined",
                           style={"fontSize": "16px", "verticalAlign": "middle", "marginRight": "6px"}),
                 "Template Notes"],
                id="btn-download-notes", n_clicks=0,
                style={
                    "background": "#f0fdf4", "color": VERT,
                    "border": f"1.5px solid {VERT}", "borderRadius": "8px",
                    "padding": "8px 14px", "fontFamily": "'Inter', sans-serif",
                    "fontWeight": "600", "fontSize": "0.82rem", "cursor": "pointer"
                }
            ),
            dcc.Upload(
                id="notes-upload",
                children=html.Button(
                    [html.Span("upload", className="material-symbols-outlined",
                               style={"fontSize": "16px", "verticalAlign": "middle", "marginRight": "6px"}),
                     "Importer Notes"],
                    style={
                        "background": "#eff6ff", "color": BLEU,
                        "border": f"1.5px solid {BLEU}", "borderRadius": "8px",
                        "padding": "8px 14px", "fontFamily": "'Inter', sans-serif",
                        "fontWeight": "600", "fontSize": "0.82rem", "cursor": "pointer"
                    }
                ),
                accept=".xlsx"
            ),
            btn_primary("+ Ajouter etudiant", "btn-open-etudiant", BLEU),
        ])
    ]),

    # -- Feedback import --
    html.Div(id="notes-feedback", style={"marginBottom": "12px"}),

    # -- Contenu principal --
    dbc.Row(style={"alignItems": "stretch"}, children=[

        # -- Liste etudiants --
        dbc.Col(
            html.Div(style={
                "background": "#ffffff", "borderRadius": "12px",
                "padding": "20px", "border": "1px solid #e5e7eb",
                "height": "100%", "display": "flex", "flexDirection": "column"
            }, children=[
                html.Div(style={
                    "display": "flex", "justifyContent": "space-between",
                    "alignItems": "center", "marginBottom": "14px"
                }, children=[
                    html.H6("Liste des etudiants", style={
                        "fontFamily": "'Montserrat', sans-serif",
                        "fontWeight": "700", "color": BLEU, "margin": "0"
                    }),
                    html.Span(id="etudiants-nb", style={
                        "background": f"{BLEU}10", "color": BLEU,
                        "padding": "2px 10px", "borderRadius": "999px",
                        "fontSize": "0.72rem", "fontWeight": "700",
                        "fontFamily": "'Montserrat', sans-serif"
                    })
                ]),
                html.Div(id="etudiants-liste",
                         style={"flex": "1", "overflowY": "auto", "maxHeight": "560px"})
            ]),
            md=5
        ),

        # -- Fiche etudiant --
        dbc.Col(
            html.Div(style={
                "background": "#ffffff", "borderRadius": "12px",
                "padding": "20px", "border": "1px solid #e5e7eb",
                "height": "100%", "display": "flex", "flexDirection": "column"
            }, children=[
                html.H6("Fiche etudiant", style={
                    "fontFamily": "'Montserrat', sans-serif",
                    "fontWeight": "700", "color": BLEU,
                    "marginBottom": "14px"
                }),
                html.Div(id="etudiant-fiche",
                         style={"flex": "1", "overflowY": "auto", "maxHeight": "560px"})
            ]),
            md=7
        )
    ], className="g-3"),

    # -- Modal ajout etudiant --
    dbc.Modal([
        dbc.ModalHeader(dbc.ModalTitle("Ajouter un etudiant")),
        dbc.ModalBody([
            dbc.Row([
                dbc.Col(field("Nom", dbc.Input(id="etud-nom", placeholder="ex: DIALLO", style=input_style())), md=6),
                dbc.Col(field("Prenom", dbc.Input(id="etud-prenom", placeholder="ex: Amadou", style=input_style())), md=6),
            ]),
            field("Email", dbc.Input(id="etud-email", type="email",
                                     placeholder="ex: adiallo@ensae.sn", style=input_style())),
            field("Matricule", dbc.Input(id="etud-matricule",
                                          placeholder="ex: ENSAE-2024-001", style=input_style())),
            field("Date de naissance", dbc.Input(id="etud-naissance", type="date", style=input_style())),
            field("Classe", dcc.Dropdown(id="etud-classe",
                                          placeholder="Selectionner une classe",
                                          style={"fontFamily": "'Inter', sans-serif", "fontSize": "0.875rem"})),
            field("Filiere d'origine", dbc.Input(id="etud-filiere-origine",
                                                   placeholder="ex: ISE Math", style=input_style())),
            field("Mot de passe provisoire", dbc.Input(id="etud-password", type="password",
                                                        placeholder="Minimum 6 caracteres", style=input_style())),
            html.Div(id="etud-feedback", style={"color": "#ef4444", "fontSize": "0.8rem"})
        ]),
        dbc.ModalFooter([
            btn_primary("Enregistrer", "btn-save-etudiant", BLEU),
            btn_outline("Annuler", "btn-cancel-etudiant")
        ])
    ], id="modal-etudiant", is_open=False, size="lg"),
])


# ============================================================
#  CALLBACKS
# ============================================================

@callback(
    Output("etudiants-filtre-classe", "options"),
    Output("notes-classe-select",     "options"),
    Output("etud-classe",             "options"),
    Input("session-store", "data")
)
def load_classes(_):
    opts = get_classes()
    return opts, opts, opts


@callback(
    Output("etudiants-liste", "children"),
    Output("etudiants-nb",    "children"),
    Input("etudiants-filtre-classe", "value"),
    Input("etudiants-search",        "value"),
    Input("etudiants-refresh",       "data")
)
def afficher_etudiants(classe_id, search, _):
    etudiants = get_etudiants(classe_id, search)
    if not etudiants:
        return html.P("Aucun etudiant trouve.", style={
            "color": "#9ca3af", "fontFamily": "'Inter', sans-serif",
            "fontSize": "0.875rem", "textAlign": "center", "padding": "20px"
        }), "0"

    items = [
        html.Div(
            id={"type": "etudiant-item", "index": e["id"]},
            style={
                "display": "grid",
                "gridTemplateColumns": "auto 1fr auto",
                "alignItems": "center",
                "gap": "12px",
                "padding": "10px 14px",
                "borderRadius": "8px",
                "border": "1px solid #e5e7eb",
                "marginBottom": "6px",
                "background": "#fafafa",
                "cursor": "pointer",
                "transition": "border-color 0.2s"
            },
            children=[
                # Avatar initiales
                html.Div(
                    f"{e['prenom'][0]}{e['nom'][0]}".upper() if e['prenom'] and e['nom'] else "?",
                    style={
                        "width": "36px", "height": "36px",
                        "borderRadius": "50%",
                        "background": f"{BLEU}15",
                        "color": BLEU, "fontWeight": "700",
                        "fontSize": "0.82rem", "fontFamily": "'Montserrat', sans-serif",
                        "display": "flex", "alignItems": "center", "justifyContent": "center",
                        "flexShrink": "0"
                    }
                ),
                # Nom + infos
                html.Div([
                    html.Span(f"{e['prenom']} {e['nom']}", style={
                        "fontWeight": "600", "color": "#111827",
                        "fontSize": "0.875rem", "fontFamily": "'Inter', sans-serif"
                    }),
                    html.Div([
                        html.Span(e["matricule"], style={
                            "color": "#9ca3af", "fontSize": "0.7rem",
                            "fontFamily": "'Inter', sans-serif", "marginRight": "8px"
                        }),
                        html.Span(e["classe"], style={
                            "color": BLEU, "fontSize": "0.7rem",
                            "fontFamily": "'Inter', sans-serif", "fontWeight": "500"
                        }),
                    ])
                ]),
                # Badges
                html.Div(style={"display": "flex", "flexDirection": "column",
                                "gap": "4px", "alignItems": "flex-end"}, children=[
                    moy_badge(e["moyenne"]),
                    taux_badge(e["taux_assiduite"])
                ])
            ]
        )
        for e in etudiants
    ]
    return html.Div(items), str(len(etudiants))


@callback(
    Output("etudiant-fiche", "children"),
    Input({"type": "etudiant-item", "index": dash.ALL}, "n_clicks"),
    prevent_initial_call=True
)
def afficher_fiche(n_clicks):
    if not any(n_clicks):
        return html.P("Selectionnez un etudiant.", style={
            "color": "#9ca3af", "fontFamily": "'Inter', sans-serif",
            "fontSize": "0.875rem", "textAlign": "center", "padding": "20px"
        })
    etudiant_id = ctx.triggered_id["index"]
    detail = get_etudiant_detail(etudiant_id)
    if not detail:
        return html.P("Etudiant introuvable.")

    notes = detail["notes"]

    # Regrouper par UE
    ues = {}
    for n in notes:
        ue = n["ue"]
        if ue not in ues:
            ues[ue] = []
        ues[ue].append(n)

    fiche = html.Div([
        # -- Entete fiche --
        html.Div(style={
            "background": f"{BLEU}08", "borderRadius": "10px",
            "padding": "16px", "marginBottom": "16px",
            "border": f"1px solid {BLEU}20"
        }, children=[
            html.Div(style={"display": "flex", "gap": "14px", "alignItems": "center"}, children=[
                html.Div(
                    f"{detail['prenom'][0]}{detail['nom'][0]}".upper(),
                    style={
                        "width": "52px", "height": "52px", "borderRadius": "50%",
                        "background": BLEU, "color": "#fff",
                        "fontFamily": "'Montserrat', sans-serif", "fontWeight": "800",
                        "fontSize": "1.1rem", "display": "flex",
                        "alignItems": "center", "justifyContent": "center", "flexShrink": "0"
                    }
                ),
                html.Div([
                    html.H6(f"{detail['prenom']} {detail['nom']}", style={
                        "fontFamily": "'Montserrat', sans-serif", "fontWeight": "800",
                        "color": BLEU, "margin": "0"
                    }),
                    html.Span(detail["matricule"], style={
                        "color": "#6b7280", "fontSize": "0.78rem",
                        "fontFamily": "'Inter', sans-serif"
                    }),
                    html.Br(),
                    html.Span(detail["classe"], style={
                        "background": f"{BLEU}15", "color": BLEU,
                        "padding": "1px 8px", "borderRadius": "4px",
                        "fontSize": "0.7rem", "fontWeight": "600",
                        "fontFamily": "'Inter', sans-serif"
                    }),
                ])
            ]),
            html.Hr(style={"borderColor": "#e5e7eb", "margin": "12px 0"}),
            dbc.Row([
                dbc.Col([
                    html.P("Assiduite", style={"color": "#9ca3af", "fontSize": "0.7rem",
                           "fontFamily": "'Inter', sans-serif", "margin": "0"}),
                    html.P(f"{100 - detail['absences']}%", style={
                        "fontWeight": "700", "color": VERT if detail['absences'] == 0 else OR,
                        "fontSize": "1rem", "fontFamily": "'Montserrat', sans-serif", "margin": "0"
                    })
                ], md=4),
                dbc.Col([
                    html.P("Absences", style={"color": "#9ca3af", "fontSize": "0.7rem",
                           "fontFamily": "'Inter', sans-serif", "margin": "0"}),
                    html.P(str(detail["absences"]), style={
                        "fontWeight": "700",
                        "color": "#ef4444" if detail["absences"] > 0 else VERT,
                        "fontSize": "1rem", "fontFamily": "'Montserrat', sans-serif", "margin": "0"
                    })
                ], md=4),
                dbc.Col([
                    html.P("Email", style={"color": "#9ca3af", "fontSize": "0.7rem",
                           "fontFamily": "'Inter', sans-serif", "margin": "0"}),
                    html.P(detail["email"], style={
                        "fontWeight": "500", "color": "#374151",
                        "fontSize": "0.78rem", "fontFamily": "'Inter', sans-serif", "margin": "0"
                    })
                ], md=4),
            ])
        ]),

        # -- Notes par UE --
        html.H6("Releve de notes", style={
            "fontFamily": "'Montserrat', sans-serif", "fontWeight": "700",
            "color": BLEU, "marginBottom": "10px"
        }),

        html.Div([
            html.Div(style={"marginBottom": "14px"}, children=[
                # Titre UE
                html.Div(ue_name, style={
                    "background": f"{VERT}10", "color": VERT,
                    "padding": "6px 12px", "borderRadius": "6px",
                    "fontWeight": "700", "fontSize": "0.78rem",
                    "fontFamily": "'Montserrat', sans-serif",
                    "marginBottom": "6px"
                }),
                # Modules de l'UE
                html.Div([
                    html.Div(style={
                        "display": "grid",
                        "gridTemplateColumns": "1fr 70px 70px 80px",
                        "gap": "8px", "alignItems": "center",
                        "padding": "8px 10px",
                        "borderRadius": "6px",
                        "border": "1px solid #f3f4f6",
                        "marginBottom": "4px",
                        "background": "#ffffff"
                    }, children=[
                        html.Span(n["module"], style={
                            "fontSize": "0.82rem", "color": "#374151",
                            "fontFamily": "'Inter', sans-serif", "fontWeight": "500"
                        }),
                        html.Span(
                            f"{n['note1']}/20" if n['note1'] is not None else "-",
                            style={"fontSize": "0.78rem", "color": "#6b7280",
                                   "fontFamily": "'Inter', sans-serif", "textAlign": "center"}
                        ),
                        html.Span(
                            f"{n['note2']}/20" if n['note2'] is not None else "-",
                            style={"fontSize": "0.78rem", "color": "#6b7280",
                                   "fontFamily": "'Inter', sans-serif", "textAlign": "center"}
                        ),
                        html.Span(
                            f"{n['moyenne']}/20" if n['moyenne'] is not None else "-",
                            style={
                                "fontSize": "0.82rem", "fontWeight": "700", "textAlign": "center",
                                "color": VERT if n['moyenne'] and n['moyenne'] >= 10 else "#ef4444",
                                "fontFamily": "'Montserrat', sans-serif"
                            }
                        )
                    ])
                    for n in ue_mods
                ]),
            ])
            for ue_name, ue_mods in ues.items()
        ]) if notes else html.P("Aucune note enregistree.", style={
            "color": "#9ca3af", "fontFamily": "'Inter', sans-serif",
            "fontSize": "0.875rem", "textAlign": "center", "padding": "20px"
        })
    ])
    return fiche


# -- Modal etudiant --
@callback(
    Output("modal-etudiant", "is_open"),
    Input("btn-open-etudiant",   "n_clicks"),
    Input("btn-cancel-etudiant", "n_clicks"),
    Input("btn-save-etudiant",   "n_clicks"),
    prevent_initial_call=True
)
def toggle_modal(o, c, s):
    return ctx.triggered_id == "btn-open-etudiant"


@callback(
    Output("etud-feedback",    "children"),
    Output("etudiants-refresh","data"),
    Output("modal-etudiant",   "is_open", allow_duplicate=True),
    Input("btn-save-etudiant", "n_clicks"),
    State("etud-nom",           "value"),
    State("etud-prenom",        "value"),
    State("etud-email",         "value"),
    State("etud-matricule",     "value"),
    State("etud-naissance",     "value"),
    State("etud-classe",        "value"),
    State("etud-filiere-origine","value"),
    State("etud-password",      "value"),
    prevent_initial_call=True
)
def save_etudiant(n, nom, prenom, email, matricule, naissance, classe_id, filiere_origine, password):
    if not all([nom, prenom, email, matricule, classe_id, password]):
        return "Tous les champs obligatoires doivent etre remplis.", None, True
    from auth import hash_password
    from datetime import datetime
    db = SessionLocal()
    try:
        if db.query(User).filter(User.email == email).first():
            return f"Email '{email}' deja utilise.", None, True
        if db.query(Etudiant).filter(Etudiant.matricule == matricule).first():
            return f"Matricule '{matricule}' deja utilise.", None, True
        user = User(
            nom=nom.upper(), prenom=prenom,
            email=email, password_hash=hash_password(password),
            role=RoleEnum.eleve, is_active=True
        )
        db.add(user)
        db.flush()
        date_naissance = datetime.strptime(naissance, "%Y-%m-%d").date() if naissance else None
        etudiant = Etudiant(
            user_id=user.id, matricule=matricule,
            date_naissance=date_naissance, classe_id=classe_id,
            filiere_origine=filiere_origine or "",
            annee_scolaire="2024-2025"
        )
        db.add(etudiant)
        db.commit()
        return "", True, False
    except Exception as e:
        db.rollback()
        return f"Erreur : {str(e)}", None, True
    finally:
        db.close()


# -- Template notes --
@callback(
    Output("download-notes-template", "data"),
    Input("btn-download-notes",        "n_clicks"),
    State("notes-classe-select",       "value"),
    prevent_initial_call=True
)
def download_template(n, classe_id):
    if not classe_id:
        return None
    data = generate_notes_template(classe_id)
    return dcc.send_bytes(data, "template_notes_ensae.xlsx")


# -- Import notes --
@callback(
    Output("notes-feedback",   "children"),
    Output("etudiants-refresh","data", allow_duplicate=True),
    Input("notes-upload",      "contents"),
    prevent_initial_call=True
)
def import_notes(contents):
    if not contents:
        return html.Div(), None
    try:
        nb, erreurs = import_notes_from_excel(contents)
        messages = []
        if nb:
            messages.append(dbc.Alert(
                f"{nb} note(s) importee(s) avec succes.",
                color="success", dismissable=True,
                style={"fontFamily": "'Inter', sans-serif", "fontSize": "0.875rem"}
            ))
        if erreurs:
            messages.append(dbc.Alert(
                [html.Strong("Avertissements :"), html.Ul([html.Li(e) for e in erreurs])],
                color="warning", dismissable=True,
                style={"fontFamily": "'Inter', sans-serif", "fontSize": "0.82rem"}
            ))
        return html.Div(messages), True
    except Exception as e:
        return dbc.Alert(f"Erreur : {str(e)}", color="danger", dismissable=True,
                         style={"fontFamily": "'Inter', sans-serif"}), None