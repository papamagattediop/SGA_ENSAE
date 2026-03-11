# ============================================================
#  SGA ENSAE — pages/cours.py
#  Gestion des UE et Modules — CRUD + Import Excel en masse
#  Python 3.11 · Dash 2.17.0
# ============================================================

import dash
from dash import html, dcc, callback, Input, Output, State, ctx
import dash_bootstrap_components as dbc
from auth import require_auth
from database import SessionLocal
from models import UE, Module, Classe, Periode, Seance, UEClasse
from utils.access_helpers import get_classes_for_user, get_default_classe_id
import pandas as pd
import io
import base64

dash.register_page(__name__, path="/cours", title="SGA ENSAE — Cours et UE")

BLEU = "#003580"
VERT = "#006B3F"
OR   = "#F5A623"


# ============================================================
#  UTILITAIRES DB
# ============================================================

def get_classes():
    db = SessionLocal()
    try:
        return [{"label": c.nom, "value": c.id} for c in db.query(Classe).all()]
    finally:
        db.close()

def get_periodes():
    db = SessionLocal()
    try:
        return [{"label": p.libelle, "value": p.id} for p in db.query(Periode).all()]
    finally:
        db.close()

def get_ues(classe_id=None):
    db = SessionLocal()
    try:
        q = db.query(UE)
        if classe_id:
            q = q.join(UE.ue_classes).filter_by(classe_id=classe_id)
        return [
            {
                "id"         : u.id,
                "code"       : u.code,
                "libelle"    : u.libelle,
                "coefficient": u.coefficient,
                "periode"    : u.periode.libelle if u.periode else "-",
                "nb_modules" : len(u.modules),
            }
            for u in q.all()
        ]
    finally:
        db.close()

def get_modules(ue_id=None):
    db = SessionLocal()
    try:
        q = db.query(Module)
        if ue_id:
            q = q.filter(Module.ue_id == ue_id)
        result = []
        for m in q.all():
            heures_faites = sum(
                int((s.heure_fin.hour * 60 + s.heure_fin.minute -
                     s.heure_debut.hour * 60 - s.heure_debut.minute) / 60)
                for s in db.query(Seance).filter(Seance.module_id == m.id).all()
                if s.heure_debut and s.heure_fin
            )
            result.append({
                "id"              : m.id,
                "code"            : m.code,
                "libelle"         : m.libelle,
                "enseignant"      : m.enseignant or "-",
                "email_enseignant": m.email_enseignant or "",
                "coefficient"     : m.coefficient,
                "volume_horaire"  : m.volume_horaire or 0,
                "heures_faites"   : heures_faites,
                "progression"     : min(100, int(heures_faites / m.volume_horaire * 100)) if m.volume_horaire else 0,
                "ue_code"         : m.ue.code if m.ue else "-",
            })
        return result
    finally:
        db.close()


# ============================================================
#  GENERATION TEMPLATE EXCEL
# ============================================================

def generate_template_excel() -> bytes:
    """
    Génère un template Excel avec trois feuilles :
    - Feuille 1 : UE      (Code_UE, Libelle_UE, Coefficient_UE, Code_Periode, Classe_ID)
    - Feuille 2 : Modules (Code_Module, Libelle_Module, Code_UE_Parent,
                           Enseignant, Email_Enseignant, Coefficient_Module,
                           Volume_Horaire_h, Classe_ID)
    - Feuille 3 : Référence_Classes
                  (Classe_ID, Nom_Classe, Code_Filiere, Niveau, Annee_Scolaire)

    IMPORTANT :
      • Classe_ID doit correspondre à l'ID affiché dans la feuille "Référence_Classes".
      • Code_Periode doit correspondre au libellé exact d'une période existante en base.
      • Email_Enseignant est utilisé pour les notifications planning — renseigner si possible.
    """
    import io
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    from database import SessionLocal
    from models import Periode, Classe, Filiere

    db = SessionLocal()
    try:
        periodes = db.query(Periode).all()
        classes  = (
            db.query(Classe)
              .join(Filiere, Classe.filiere_id == Filiere.id)
              .order_by(Filiere.code, Classe.niveau, Classe.nom)
              .all()
        )

        exemple_classe_id   = classes[0].id      if classes  else 1
        exemple_periode_lib = periodes[0].libelle if periodes else "Semestre 1"

        # ── Référence classes ──────────────────────────────────
        ref_data = []
        for c in classes:
            ref_data.append({
                "Classe_ID"      : c.id,
                "Nom_Classe"     : c.nom,
                "Code_Filiere"   : c.filiere.code   if c.filiere else "",
                "Filiere_Libelle": c.filiere.libelle if c.filiere else "",
                "Niveau"         : c.niveau,
                "Annee_Scolaire" : c.annee_scolaire,
            })
        df_ref = pd.DataFrame(ref_data) if ref_data else pd.DataFrame(columns=[
            "Classe_ID", "Nom_Classe", "Code_Filiere", "Filiere_Libelle",
            "Niveau", "Annee_Scolaire"
        ])

    finally:
        db.close()

    # ── DataFrame UE ──────────────────────────────────────────
    df_ue = pd.DataFrame(columns=[
        "Code_UE", "Libelle_UE", "Coefficient_UE", "Code_Periode", "Classe_ID"
    ])
    df_ue.loc[0] = ["UE-STAT", "Statistiques et Probabilites", 4, exemple_periode_lib, exemple_classe_id]
    df_ue.loc[1] = ["UE-ECO",  "Economie Generale",            3, exemple_periode_lib, exemple_classe_id]

    # ── DataFrame Modules — MODIFIÉ : ajout Email_Enseignant ──
    df_mod = pd.DataFrame(columns=[
        "Code_Module", "Libelle_Module", "Code_UE_Parent",
        "Enseignant", "Email_Enseignant",
        "Coefficient_Module", "Volume_Horaire_h", "Classe_ID"
    ])
    df_mod.loc[0] = ["MOD-STAT1", "Statistiques Descriptives", "UE-STAT",
                     "Dr. Diallo", "diallo@ensae.sn",  2, 30, exemple_classe_id]
    df_mod.loc[1] = ["MOD-STAT2", "Probabilites",              "UE-STAT",
                     "Dr. Ndiaye", "ndiaye@ensae.sn",  2, 25, exemple_classe_id]
    df_mod.loc[2] = ["MOD-ECO1",  "Microeconomie",             "UE-ECO",
                     "Dr. Ba",     "ba@ensae.sn",       3, 40, exemple_classe_id]

    # ── Écriture Excel ─────────────────────────────────────────
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_ue.to_excel(writer,  sheet_name="UE",                index=False)
        df_mod.to_excel(writer, sheet_name="Modules",           index=False)
        df_ref.to_excel(writer, sheet_name="Référence_Classes", index=False)

        wb = writer.book
        from openpyxl.styles import PatternFill, Font, Alignment

        FILL_UE  = PatternFill("solid", fgColor="003580")
        FILL_MOD = PatternFill("solid", fgColor="006B3F")
        FILL_REF = PatternFill("solid", fgColor="4B0082")
        FILL_CID = PatternFill("solid", fgColor="FFF3CD")
        FILL_EMAIL = PatternFill("solid", fgColor="E8F5E9")   # vert pâle — Email_Enseignant
        FONT_W   = Font(color="FFFFFF", bold=True, name="Calibri")
        ALIGN_C  = Alignment(horizontal="center")

        # ── Style feuille UE ───────────────────────────────────
        ws_ue = writer.sheets["UE"]
        for cell in ws_ue[1]:
            cell.fill      = FILL_UE
            cell.font      = FONT_W
            cell.alignment = ALIGN_C
        for col in ws_ue.columns:
            ws_ue.column_dimensions[col[0].column_letter].width = 26
        _highlight_classe_id_col(ws_ue, col_letter="E", fill=FILL_CID)
        ws_ue["E1"].comment = _make_comment(
            "⚠ Classe_ID requis\nConsultez la feuille 'Référence_Classes' "
            "pour obtenir l'ID de votre classe."
        )

        nb_classes = len(df_ref)
        if nb_classes > 0:
            dv_ue = DataValidation(
                type="list",
                formula1=f"'Référence_Classes'!$A$2:$A${nb_classes + 1}",
                allow_blank=False,
                showErrorMessage=True,
                errorTitle="Classe_ID invalide",
                error="Choisissez un ID dans la feuille 'Référence_Classes'."
            )
            ws_ue.add_data_validation(dv_ue)
            dv_ue.add("E2:E1000")

        # ── Style feuille Modules ──────────────────────────────
        # MODIFIÉ : Email_Enseignant est en col E (nouveau), Classe_ID en col H
        ws_mod = writer.sheets["Modules"]
        for cell in ws_mod[1]:
            cell.fill      = FILL_MOD
            cell.font      = FONT_W
            cell.alignment = ALIGN_C
        for col in ws_mod.columns:
            ws_mod.column_dimensions[col[0].column_letter].width = 26

        # Mettre en vert pâle la colonne Email_Enseignant (col E = 5)
        _highlight_classe_id_col(ws_mod, col_letter="E", fill=FILL_EMAIL)
        ws_mod["E1"].comment = _make_comment(
            "Email de l'enseignant (optionnel mais recommandé).\n"
            "Utilisé pour les notifications automatiques lors de la validation du planning."
        )

        # Colonne Classe_ID = col H (décalée car Email_Enseignant ajouté)
        _highlight_classe_id_col(ws_mod, col_letter="H", fill=FILL_CID)
        ws_mod["H1"].comment = _make_comment(
            "⚠ Classe_ID requis\nConsultez la feuille 'Référence_Classes' "
            "pour obtenir l'ID de votre classe."
        )
        if nb_classes > 0:
            dv_mod = DataValidation(
                type="list",
                formula1=f"'Référence_Classes'!$A$2:$A${nb_classes + 1}",
                allow_blank=False,
                showErrorMessage=True,
                errorTitle="Classe_ID invalide",
                error="Choisissez un ID dans la feuille 'Référence_Classes'."
            )
            ws_mod.add_data_validation(dv_mod)
            dv_mod.add("H2:H1000")

        # ── Style feuille Référence_Classes ───────────────────
        ws_ref = writer.sheets["Référence_Classes"]
        for cell in ws_ref[1]:
            cell.fill      = FILL_REF
            cell.font      = FONT_W
            cell.alignment = ALIGN_C

        FILL_ID_REF = PatternFill("solid", fgColor="DBEAFE")
        for row in ws_ref.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                cell.fill = FILL_ID_REF
                cell.font = Font(bold=True, name="Calibri", color="003580")
                cell.alignment = ALIGN_C

        col_widths = [12, 22, 16, 38, 10, 14]
        for i, w in enumerate(col_widths, start=1):
            ws_ref.column_dimensions[get_column_letter(i)].width = w

        ws_ref.freeze_panes = "A2"
        ws_ref["A1"].comment = _make_comment(
            "⚠ NE PAS MODIFIER cette feuille.\n"
            "Elle est générée automatiquement depuis la base de données.\n"
            "Utilisez les IDs de la colonne 'Classe_ID' dans les feuilles UE et Modules."
        )

    buffer.seek(0)
    return buffer.read()


# ── Helpers ───────────────────────────────────────────────────

def _highlight_classe_id_col(ws, col_letter: str, fill):
    """Colore le fond de toutes les cellules de la colonne (hors en-tête)."""
    from openpyxl.styles import Font
    for row in ws.iter_rows(min_row=2, min_col=ord(col_letter) - 64,
                             max_col=ord(col_letter) - 64, max_row=100):
        for cell in row:
            cell.fill = fill


def _make_comment(text: str):
    """Crée un commentaire openpyxl."""
    from openpyxl.comments import Comment
    c = Comment(text, "SGA ENSAE")
    c.width  = 280
    c.height = 90
    return c


# ============================================================
#  IMPORT EXCEL EN MASSE
# ============================================================

def import_from_excel(contents: str) -> tuple[int, int, list]:
    """
    Importe UE et Modules depuis un fichier Excel.
    Retourne (nb_ue, nb_modules, erreurs/avertissements).

    MODIFIÉ : lecture de Email_Enseignant dans la feuille Modules.
    La recherche de periode se fait par (libelle + classe_id) avec
    fallback sur libelle seul pour compatibilité.
    """
    content_type, content_string = contents.split(",")
    decoded   = base64.b64decode(content_string)
    xl        = pd.ExcelFile(io.BytesIO(decoded))

    erreurs    = []
    nb_ue      = 0
    nb_modules = 0
    db         = SessionLocal()

    try:
        # ── Import UE ──────────────────────────────────────────
        if "UE" in xl.sheet_names:
            df_ue = xl.parse("UE")
            for idx, row in df_ue.iterrows():
                try:
                    code = str(row["Code_UE"]).strip()
                    if not code or code.upper() == "NAN":
                        continue

                    if db.query(UE).filter(UE.code == code).first():
                        erreurs.append(f"UE '{code}' existe deja — ignoree.")
                        continue

                    libelle_periode = str(row["Code_Periode"]).strip()
                    classe_id       = int(row["Classe_ID"])

                    # Recherche periode par (libelle + classe_id) puis fallback libelle seul
                    periode = db.query(Periode).filter(
                        Periode.libelle   == libelle_periode,
                        Periode.classe_id == classe_id
                    ).first()
                    if not periode:
                        periode = db.query(Periode).filter(
                            Periode.libelle == libelle_periode
                        ).first()
                    if not periode:
                        erreurs.append(
                            f"Periode '{libelle_periode}' introuvable pour UE '{code}' "
                            f"(classe_id={classe_id})."
                        )
                        continue

                    coef = float(row.get("Coefficient_UE", 1.0))
                    ue   = UE(
                        code       = code,
                        libelle    = str(row["Libelle_UE"]).strip(),
                        coefficient= coef,
                        periode_id = periode.id,
                    )
                    db.add(ue)
                    db.flush()
                    db.add(UEClasse(ue_id=ue.id, classe_id=classe_id))
                    nb_ue += 1

                except Exception as e:
                    erreurs.append(f"UE ligne {idx + 2} : {str(e)}")

        # ── Import Modules — MODIFIÉ : lecture Email_Enseignant ──
        if "Modules" in xl.sheet_names:
            df_mod = xl.parse("Modules")
            for idx, row in df_mod.iterrows():
                try:
                    code = str(row["Code_Module"]).strip()
                    if not code or code.upper() == "NAN":
                        continue

                    if db.query(Module).filter(Module.code == code).first():
                        erreurs.append(f"Module '{code}' existe deja — ignore.")
                        continue

                    code_ue   = str(row["Code_UE_Parent"]).strip()
                    classe_id = int(row["Classe_ID"])
                    ue        = db.query(UE).filter(UE.code == code_ue).first()
                    if not ue:
                        erreurs.append(
                            f"UE parente '{code_ue}' introuvable pour module '{code}'."
                        )
                        continue

                    enseignant = str(row.get("Enseignant", "")).strip()
                    if enseignant.upper() == "NAN":
                        enseignant = ""

                    # ← NOUVEAU : lecture Email_Enseignant
                    email_ens = str(row.get("Email_Enseignant", "")).strip()
                    if email_ens.upper() == "NAN":
                        email_ens = ""

                    m = Module(
                        code             = code,
                        libelle          = str(row["Libelle_Module"]).strip(),
                        enseignant       = enseignant or None,
                        email_enseignant = email_ens or None,   # ← NOUVEAU
                        coefficient      = float(row.get("Coefficient_Module", 1.0)),
                        volume_horaire   = int(row.get("Volume_Horaire_h", 0)),
                        ue_id            = ue.id,
                        classe_id        = classe_id,
                    )
                    db.add(m)
                    nb_modules += 1

                except Exception as e:
                    erreurs.append(f"Module ligne {idx + 2} : {str(e)}")

        db.commit()

    except Exception as e:
        db.rollback()
        erreurs.append(f"Erreur globale : {str(e)}")
    finally:
        db.close()

    return nb_ue, nb_modules, erreurs


# ============================================================
#  LAYOUT
# ============================================================

def input_style():
    return {
        "fontFamily": "'Inter', sans-serif",
        "fontSize"  : "0.875rem",
        "borderRadius": "8px",
        "border"    : "1px solid #e5e7eb",
    }

def field(label, component):
    return html.Div([
        html.Label(label, style={
            "fontFamily": "'Inter', sans-serif",
            "fontSize"  : "0.78rem",
            "fontWeight": "600",
            "color"     : "#374151",
            "marginBottom": "4px",
            "display"   : "block"
        }),
        component
    ], style={"marginBottom": "14px"})

def btn_primary(label, id_, color):
    return html.Button(label, id=id_, n_clicks=0, style={
        "background"  : color,
        "color"       : "#ffffff",
        "border"      : "none",
        "borderRadius": "8px",
        "padding"     : "8px 20px",
        "fontFamily"  : "'Inter', sans-serif",
        "fontWeight"  : "600",
        "fontSize"    : "0.82rem",
        "cursor"      : "pointer",
        "marginRight" : "8px",
    })

def btn_outline(label, id_):
    return html.Button(label, id=id_, n_clicks=0, style={
        "background"  : "transparent",
        "color"       : "#6b7280",
        "border"      : "1px solid #d1d5db",
        "borderRadius": "8px",
        "padding"     : "8px 20px",
        "fontFamily"  : "'Inter', sans-serif",
        "fontWeight"  : "500",
        "fontSize"    : "0.82rem",
        "cursor"      : "pointer",
    })

def progress_bar(pct):
    color = VERT if pct >= 75 else OR if pct >= 40 else "#ef4444"
    return html.Div([
        html.Div(style={
            "width": f"{pct}%", "height": "6px",
            "background": color, "borderRadius": "4px",
            "transition": "width 0.3s ease"
        })
    ], style={
        "width": "100%", "height": "6px",
        "background": "#e5e7eb", "borderRadius": "4px"
    })


layout = html.Div([
    dcc.Store(id="session-store", storage_type="session"),
    dcc.Download(id="download-template-excel"),

    # -- Retour + En-tête --
    html.Div(style={"marginBottom": "24px"}, children=[
        html.A(href="/dashboard", style={
            "textDecoration": "none", "display": "inline-flex",
            "alignItems": "center", "gap": "6px", "marginBottom": "12px",
            "color": "#6b7280", "fontFamily": "'Inter', sans-serif",
            "fontSize": "0.82rem", "fontWeight": "500"
        }, children=[
            html.Span("arrow_back", className="material-symbols-outlined",
                      style={"fontSize": "18px", "verticalAlign": "middle"}),
            "Retour au tableau de bord"
        ]),
        html.H4("Cours et UE", style={
            "fontFamily": "'Montserrat', sans-serif",
            "fontWeight": "800", "color": BLEU, "margin": "0"
        }),
        html.P("Gestion des Unités d'Enseignement et des modules pédagogiques", style={
            "color": "#6b7280", "fontFamily": "'Inter', sans-serif", "margin": "4px 0 0"
        })
    ]),

    # -- Barre d'actions globale --
    html.Div(
        style={
            "background": "#ffffff", "borderRadius": "12px",
            "padding": "16px 24px", "boxShadow": "0 2px 12px rgba(0,0,0,0.07)",
            "marginBottom": "20px", "display": "flex",
            "alignItems": "center", "gap": "12px", "flexWrap": "wrap"
        },
        children=[
            html.Div(style={"flex": "1", "minWidth": "200px"}, children=[
                dcc.Dropdown(
                    id="cours-filtre-classe",
                    placeholder="Filtrer par classe...",
                    style={"fontFamily": "'Inter', sans-serif", "fontSize": "0.875rem"}
                )
            ]),
            html.Button(
                [
                    html.Span("download", className="material-symbols-outlined",
                              style={"fontSize": "16px", "verticalAlign": "middle", "marginRight": "6px"}),
                    "Template Excel"
                ],
                id="btn-download-template", n_clicks=0,
                style={
                    "background": "#f0fdf4", "color": VERT,
                    "border": f"1.5px solid {VERT}", "borderRadius": "8px",
                    "padding": "8px 16px", "fontFamily": "'Inter', sans-serif",
                    "fontWeight": "600", "fontSize": "0.82rem", "cursor": "pointer"
                }
            ),
            dcc.Upload(
                id="cours-upload-excel",
                children=html.Button(
                    [
                        html.Span("upload", className="material-symbols-outlined",
                                  style={"fontSize": "16px", "verticalAlign": "middle", "marginRight": "6px"}),
                        "Importer Excel"
                    ],
                    style={
                        "background": "#eff6ff", "color": BLEU,
                        "border": f"1.5px solid {BLEU}", "borderRadius": "8px",
                        "padding": "8px 16px", "fontFamily": "'Inter', sans-serif",
                        "fontWeight": "600", "fontSize": "0.82rem", "cursor": "pointer"
                    }
                ),
                accept=".xlsx"
            ),
            btn_primary("+ Ajouter UE",     "btn-open-ue",     BLEU),
            btn_primary("+ Ajouter Module", "btn-open-module", VERT),
        ]
    ),

    # -- Feedback import --
    html.Div(id="import-feedback", style={"marginBottom": "16px"}),

    # -- Colonnes UE | Modules --
    dbc.Row(
        style={"alignItems": "stretch"},
        children=[
            # Colonne UE
            dbc.Col(
                html.Div(
                    style={
                        "background": "#ffffff", "borderRadius": "12px",
                        "padding": "20px", "boxShadow": "0 2px 12px rgba(0,0,0,0.07)",
                        "height": "100%", "display": "flex", "flexDirection": "column"
                    },
                    children=[
                        html.Div(style={
                            "display": "flex", "justifyContent": "space-between",
                            "alignItems": "center", "marginBottom": "14px"
                        }, children=[
                            html.H6("Unites d'Enseignement", style={
                                "fontFamily": "'Montserrat', sans-serif",
                                "fontWeight": "700", "color": BLEU, "margin": "0"
                            }),
                            html.Span(id="cours-nb-ue", style={
                                "background": f"{BLEU}15", "color": BLEU,
                                "padding": "2px 10px", "borderRadius": "999px",
                                "fontSize": "0.72rem", "fontWeight": "700",
                                "fontFamily": "'Montserrat', sans-serif"
                            })
                        ]),
                        html.Div(
                            id="cours-liste-ue",
                            style={"flex": "1", "overflowY": "auto", "maxHeight": "520px"}
                        )
                    ]
                ),
                md=5
            ),

            # Colonne Modules
            dbc.Col(
                html.Div(
                    style={
                        "background": "#ffffff", "borderRadius": "12px",
                        "padding": "20px", "boxShadow": "0 2px 12px rgba(0,0,0,0.07)",
                        "height": "100%", "display": "flex", "flexDirection": "column"
                    },
                    children=[
                        html.Div(style={
                            "display": "flex", "justifyContent": "space-between",
                            "alignItems": "center", "marginBottom": "6px"
                        }, children=[
                            html.H6("Modules", style={
                                "fontFamily": "'Montserrat', sans-serif",
                                "fontWeight": "700", "color": VERT, "margin": "0"
                            }),
                            html.Span(id="cours-nb-modules", style={
                                "background": f"{VERT}15", "color": VERT,
                                "padding": "2px 10px", "borderRadius": "999px",
                                "fontSize": "0.72rem", "fontWeight": "700",
                                "fontFamily": "'Montserrat', sans-serif"
                            })
                        ]),
                        html.P(
                            "Cliquez sur une UE pour afficher ses modules.",
                            id="cours-hint-modules",
                            style={"color": "#9ca3af", "fontSize": "0.78rem",
                                   "fontFamily": "'Inter', sans-serif", "margin": "0 0 10px"}
                        ),
                        html.Div(
                            id="cours-liste-modules",
                            style={"flex": "1", "overflowY": "auto", "maxHeight": "500px"}
                        )
                    ]
                ),
                md=7
            ),
        ],
        className="g-3"
    ),

    # -- Modal UE --
    dbc.Modal([
        dbc.ModalHeader(dbc.ModalTitle("Ajouter une UE")),
        dbc.ModalBody([
            field("Code UE",    dbc.Input(id="ue-code",    placeholder="ex: UE-STAT",  style=input_style())),
            field("Libelle",    dbc.Input(id="ue-libelle", placeholder="ex: Statistiques et Probabilites", style=input_style())),
            field("Periode",    dcc.Dropdown(id="ue-periode",  placeholder="Selectionner une periode",
                                             style={"fontFamily": "'Inter', sans-serif", "fontSize": "0.875rem"})),
            field("Classe(s)",  dcc.Dropdown(id="ue-classes",  placeholder="Selectionner une ou plusieurs classes",
                                             multi=True,
                                             style={"fontFamily": "'Inter', sans-serif", "fontSize": "0.875rem"})),
            field("Coefficient", dbc.Input(id="ue-coefficient", type="number", min=1, max=10, step=0.5, value=1, style=input_style())),
            html.Div(id="ue-feedback", style={"color": "#ef4444", "fontSize": "0.8rem"})
        ]),
        dbc.ModalFooter([
            btn_primary("Enregistrer", "btn-save-ue", BLEU),
            btn_outline("Annuler",     "btn-cancel-ue")
        ])
    ], id="modal-ue", is_open=False),

    # -- Modal Module — MODIFIÉ : champ Email Enseignant ajouté --
    dbc.Modal([
        dbc.ModalHeader(dbc.ModalTitle("Ajouter un Module")),
        dbc.ModalBody([
            field("UE parente",  dcc.Dropdown(id="module-ue",    placeholder="Selectionner une UE",
                                              style={"fontFamily": "'Inter', sans-serif", "fontSize": "0.875rem"})),
            field("Classe",      dcc.Dropdown(id="module-classe", placeholder="Selectionner une classe",
                                              style={"fontFamily": "'Inter', sans-serif", "fontSize": "0.875rem"})),
            field("Code Module", dbc.Input(id="module-code",      placeholder="ex: MOD-STAT1", style=input_style())),
            field("Libelle",     dbc.Input(id="module-libelle",   placeholder="ex: Statistiques Descriptives", style=input_style())),
            field("Enseignant",  dbc.Input(id="module-enseignant", placeholder="ex: Dr. Diallo", style=input_style())),
            # ← NOUVEAU champ
            field("Email Enseignant",
                  dbc.Input(
                      id="module-email-enseignant",
                      type="email",
                      placeholder="ex: diallo@ensae.sn  (optionnel — pour les notifs planning)",
                      style=input_style()
                  )),
            field("Coefficient", dbc.Input(id="module-coefficient", type="number", min=1, max=10, step=0.5, value=1, style=input_style())),
            field("Volume horaire (h)", dbc.Input(id="module-volume", type="number", min=1, step=1, placeholder="ex: 30", style=input_style())),
            html.Div(id="module-feedback", style={"color": "#ef4444", "fontSize": "0.8rem"})
        ]),
        dbc.ModalFooter([
            btn_primary("Enregistrer", "btn-save-module", VERT),
            btn_outline("Annuler",     "btn-cancel-module")
        ])
    ], id="modal-module", is_open=False),

    html.Div(id="cours-refresh", style={"display": "none"})
])


# ============================================================
#  CALLBACKS
# ============================================================

@callback(
    Output("cours-filtre-classe", "options"),
    Output("cours-filtre-classe", "value"),
    Output("module-classe",        "options"),
    Output("ue-classes",           "options"),
    Input("session-store", "data")
)
def load_classes(session):
    if not session:
        return [], None, [], []
    role    = session.get("role", "")
    user_id = session.get("user_id")
    opts    = get_classes_for_user(role, user_id)
    default = get_default_classe_id(role, user_id)
    return opts, default, opts, opts


@callback(
    Output("ue-periode", "options"),
    Input("session-store", "data")
)
def load_periodes(_):
    return get_periodes()


@callback(
    Output("module-ue", "options"),
    Input("cours-filtre-classe", "value"),
    Input("cours-refresh", "children")
)
def load_ues_dropdown(classe_id, _):
    ues = get_ues(classe_id)
    return [{"label": f"{u['code']} — {u['libelle']}", "value": u["id"]} for u in ues]


@callback(
    Output("cours-liste-ue", "children"),
    Output("cours-nb-ue",    "children"),
    Input("cours-filtre-classe", "value"),
    Input("cours-refresh",       "children")
)
def afficher_ues(classe_id, _):
    ues = get_ues(classe_id)
    if not ues:
        return html.P("Aucune UE trouvee.", style={
            "color": "#9ca3af", "fontFamily": "'Inter', sans-serif",
            "fontSize": "0.875rem", "textAlign": "center", "padding": "20px"
        }), "0"

    items = [
        html.Div(
            id={"type": "ue-item", "index": u["id"]},
            n_clicks=0,
            style={
                "padding": "12px 14px", "borderRadius": "8px",
                "border": "1px solid #e5e7eb", "marginBottom": "8px",
                "cursor": "pointer", "background": "#fafafa"
            },
            children=[
                html.Div(style={"display": "flex", "justifyContent": "space-between"}, children=[
                    html.Span(f"{u['code']} — {u['libelle']}", style={
                        "fontFamily": "'Inter', sans-serif",
                        "fontWeight": "600", "fontSize": "0.875rem", "color": BLEU
                    }),
                    html.Span(f"Coef. {u['coefficient']}", style={
                        "color": "#6b7280", "fontSize": "0.72rem",
                        "fontFamily": "'Inter', sans-serif"
                    })
                ]),
                html.Span(f"Periode : {u['periode']}  |  {u['nb_modules']} module(s)", style={
                    "color": "#9ca3af", "fontSize": "0.72rem",
                    "fontFamily": "'Inter', sans-serif"
                })
            ]
        )
        for u in ues
    ]
    return html.Div(items), str(len(ues))


@callback(
    Output("cours-liste-modules", "children"),
    Output("cours-nb-modules",    "children"),
    Output("cours-hint-modules",  "style"),
    Input({"type": "ue-item", "index": dash.ALL}, "n_clicks"),
    Input("cours-refresh", "children"),
    prevent_initial_call=True
)
def afficher_modules(clicks, _):
    hint_hidden = {"display": "none"}
    if not any(clicks):
        return html.Div(), "0", {"color": "#9ca3af", "fontSize": "0.78rem",
                                  "fontFamily": "'Inter', sans-serif", "margin": "0 0 10px"}

    triggered = ctx.triggered_id
    if not triggered or not isinstance(triggered, dict):
        return html.Div(), "0", hint_hidden

    ue_id   = triggered["index"]
    modules = get_modules(ue_id)

    if not modules:
        return html.P("Aucun module pour cette UE.", style={
            "color": "#9ca3af", "fontFamily": "'Inter', sans-serif",
            "fontSize": "0.875rem", "textAlign": "center", "padding": "20px"
        }), "0", hint_hidden

    items = [
        html.Div(
            style={
                "padding": "12px 14px", "borderRadius": "8px",
                "border": "1px solid #e5e7eb", "marginBottom": "8px",
                "background": "#fafafa"
            },
            children=[
                html.Div(style={"display": "flex", "justifyContent": "space-between"}, children=[
                    html.Span(f"{m['code']} — {m['libelle']}", style={
                        "fontFamily": "'Inter', sans-serif",
                        "fontWeight": "600", "fontSize": "0.875rem", "color": VERT
                    }),
                    html.Span(f"Coef. {m['coefficient']}", style={
                        "color": "#6b7280", "fontSize": "0.72rem",
                        "fontFamily": "'Inter', sans-serif"
                    })
                ]),
                html.Span(
                    f"Enseignant : {m['enseignant']}"
                    + (f"  ({m['email_enseignant']})" if m['email_enseignant'] else "")
                    + f"  |  {m['heures_faites']}h / {m['volume_horaire']}h",
                    style={
                        "color": "#6b7280", "fontSize": "0.75rem",
                        "fontFamily": "'Inter', sans-serif",
                        "display": "block", "marginBottom": "8px"
                    }
                ),
                progress_bar(m["progression"])
            ]
        )
        for m in modules
    ]
    return html.Div(items), str(len(modules)), hint_hidden


# -- Modals --
@callback(
    Output("modal-ue", "is_open"),
    Input("btn-open-ue",   "n_clicks"),
    Input("btn-cancel-ue", "n_clicks"),
    Input("btn-save-ue",   "n_clicks"),
    prevent_initial_call=True
)
def toggle_modal_ue(o, c, s):
    return ctx.triggered_id == "btn-open-ue"


@callback(
    Output("modal-module", "is_open"),
    Input("btn-open-module",   "n_clicks"),
    Input("btn-cancel-module", "n_clicks"),
    Input("btn-save-module",   "n_clicks"),
    prevent_initial_call=True
)
def toggle_modal_module(o, c, s):
    return ctx.triggered_id == "btn-open-module"


# -- Sauvegarde UE --
@callback(
    Output("ue-feedback",   "children"),
    Output("cours-refresh", "children"),
    Output("modal-ue",      "is_open", allow_duplicate=True),
    Input("btn-save-ue",    "n_clicks"),
    State("ue-code",        "value"),
    State("ue-libelle",     "value"),
    State("ue-periode",     "value"),
    State("ue-classes",     "value"),
    State("ue-coefficient", "value"),
    prevent_initial_call=True
)
def save_ue(n, code, libelle, periode_id, classes, coef):
    if not all([code, libelle, periode_id, classes, coef]):
        return "Tous les champs sont obligatoires.", "", True
    db = SessionLocal()
    try:
        if db.query(UE).filter(UE.code == code).first():
            return f"Le code '{code}' existe deja.", "", True
        ue = UE(code=code, libelle=libelle, periode_id=periode_id, coefficient=float(coef))
        db.add(ue)
        db.flush()
        for cid in classes:
            db.add(UEClasse(ue_id=ue.id, classe_id=cid))
        db.commit()
        return "", "refresh", False
    except Exception as e:
        db.rollback()
        return f"Erreur : {str(e)}", "", True
    finally:
        db.close()


# -- Sauvegarde Module — MODIFIÉ : State + stockage email_enseignant --
@callback(
    Output("module-feedback",        "children"),
    Output("cours-refresh",          "children", allow_duplicate=True),
    Output("modal-module",           "is_open",  allow_duplicate=True),
    Input("btn-save-module",         "n_clicks"),
    State("module-ue",               "value"),
    State("module-classe",           "value"),
    State("module-code",             "value"),
    State("module-libelle",          "value"),
    State("module-enseignant",       "value"),
    State("module-email-enseignant", "value"),   # ← NOUVEAU
    State("module-coefficient",      "value"),
    State("module-volume",           "value"),
    prevent_initial_call=True
)
def save_module(n, ue_id, classe_id, code, libelle, enseignant, email_enseignant, coef, volume):
    if not all([ue_id, classe_id, code, libelle, coef, volume]):
        return "Tous les champs obligatoires doivent etre remplis.", "", True
    db = SessionLocal()
    try:
        if db.query(Module).filter(Module.code == code).first():
            return f"Le code '{code}' existe deja.", "", True

        # Nettoyage email
        email_ens = (email_enseignant or "").strip() or None

        m = Module(
            code             = code,
            libelle          = libelle,
            enseignant       = enseignant,
            email_enseignant = email_ens,   # ← NOUVEAU
            coefficient      = float(coef),
            volume_horaire   = int(volume),
            ue_id            = ue_id,
            classe_id        = classe_id,
        )
        db.add(m)
        db.commit()
        return "", "refresh", False
    except Exception as e:
        db.rollback()
        return f"Erreur : {str(e)}", "", True
    finally:
        db.close()


# -- Telechargement template Excel --
@callback(
    Output("download-template-excel", "data"),
    Input("btn-download-template",    "n_clicks"),
    prevent_initial_call=True
)
def download_template(n):
    data = generate_template_excel()
    return dcc.send_bytes(data, "template_ue_modules_ensae.xlsx")


# -- Import Excel --
@callback(
    Output("import-feedback",   "children"),
    Output("cours-refresh",     "children", allow_duplicate=True),
    Input("cours-upload-excel", "contents"),
    prevent_initial_call=True
)
def import_excel(contents):
    if not contents:
        return html.Div(), ""
    try:
        nb_ue, nb_mod, erreurs = import_from_excel(contents)
        messages = []
        if nb_ue or nb_mod:
            messages.append(
                dbc.Alert(
                    f"{nb_ue} UE et {nb_mod} module(s) importes avec succes.",
                    color="success", dismissable=True,
                    style={"fontFamily": "'Inter', sans-serif", "fontSize": "0.875rem"}
                )
            )
        if erreurs:
            messages.append(
                dbc.Alert(
                    [html.Strong("Avertissements :"), html.Ul([html.Li(e) for e in erreurs])],
                    color="warning", dismissable=True,
                    style={"fontFamily": "'Inter', sans-serif", "fontSize": "0.82rem"}
                )
            )
        return html.Div(messages), "refresh"
    except Exception as e:
        return dbc.Alert(
            f"Erreur lors de l'import : {str(e)}",
            color="danger", dismissable=True,
            style={"fontFamily": "'Inter', sans-serif", "fontSize": "0.875rem"}
        ), ""